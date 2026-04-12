# WhatsApp to Excel - Ultimate Fixed Version

# Cell 1: Install Packages
import subprocess
import sys
packages = ['selenium', 'openpyxl', 'webdriver-manager', 'pandas']
for package in packages:
    try:
        __import__(package.replace('-', '_'))
        print(f"✓ {package}")
    except:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])
        print(f"✓ {package} installed")
print("✅ READY!")

## 2. Imports

import os, time, json, re, hashlib
from datetime import datetime
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import pandas as pd

print("✅ Libraries loaded!")

## 3. Ultimate Automation Class

class WhatsAppExcelUltimate:
    def __init__(self, excel_path=None, group_name='Hns Surveillance'):
        self.excel_path = self.find_excel_file(excel_path)
        self.group_name = group_name
        self.driver = None
        self.processed = set()
        self.load_cache()
        print(f"📁 Excel: {self.excel_path}")
        print(f"👥 Group: {self.group_name}")

    def find_excel_file(self, path):
        """Auto-find Excel file on Desktop"""
        desktop = Path.home() / "Desktop"
        for ext in ['*.xlsx', '*.xlsm']:
            for file in desktop.glob(ext):
                if 'surveillance' in file.name.lower():
                    return str(file)
        return path or "Feb_surveillance.xlsx"

    def load_cache(self):
        cache = Path("processed_messages.json")
        if cache.exists():
            with open(cache, 'r') as f:
                self.processed = set(json.load(f))

    def save_cache(self):
        with open("processed_messages.json", 'w') as f:
            json.dump(list(self.processed), f)

    def safe_start_chrome(self):
        \"\"\"🔧 FIXED LOGIN - Longer wait + manual QR fallback\"\"\"
        options = Options()
        options.add_argument(\"--no-sandbox\")
        options.add_argument(\"--disable-dev-shm-usage\")
        options.add_argument(\"--disable-gpu\")
        options.add_argument(\"--disable-web-security\")
        options.add_argument(\"--start-maximized\")
        options.add_argument(\"--remote-debugging-port=9222\")
        
        # Try existing profile first, fallback to fresh
        profile_dirs = [
            str(Path.home() / \".whatsapp_ultimate\"),
            str(Path.home() / \"chrome-data\"),
            None  # Fresh profile
        ]
        
        for profile_dir in profile_dirs:
            try:
                if profile_dir:
                    options.add_argument(f\"--user-data-dir={profile_dir}\")
                
                service = Service(ChromeDriverManager().install())
                self.driver = webdriver.Chrome(service=service, options=options)
                self.driver.get(\"https://web.whatsapp.com\")
                
                print(\"🔐 QR CODE SHOWN - Scan with phone:\")
                print(\"📱 WhatsApp > Linked Devices > Link Device\")
                
                # Wait longer for QR + multiple checks
                for wait_time in [30, 60, 90, 120]:
                    try:
                        WebDriverWait(self.driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, \"//div[@data-testid='chat-list']\"))  
                        )
                        print(\"✅ AUTO LOGIN SUCCESS!\")
                        return True
                    except:
                        print(f\"⏳ Waiting... ({wait_time}s)\")
                        continue
                
                # Manual fallback
                print(\"\\n🎯 MANUAL LOGIN:\")
                print(\"If QR expired, press Ctrl+C and run again\")
                input(\"Press ENTER after scanning QR and seeing chats...\")
                print(\"✅ LOGIN CONFIRMED!\")
                return True
                
            except Exception as e:
                print(f\"Profile failed: {e}\")
                if self.driver:
                    self.driver.quit()
                options = Options()  # Reset options
                continue
        
        print(\"❌ All login attempts failed\")
        return False

    def find_group(self):
        """🔧 FIXED: Multiple search methods + manual fallback"""
        print(f"🔍 Searching: '{self.group_name}'")
        
        search_selectors = [
            "//input[@placeholder*='Search or start new chat']",
            "//input[@placeholder*='Search']",
            "//div[@contenteditable='true'][@data-tab='3']",
            "//input[@title*='Search']"
        ]
        
        for attempt in range(3):
            try:
                # Try different search boxes
                search_box = None
                for selector in search_selectors:
                    try:
                        search_box = WebDriverWait(self.driver, 3).until(
                            EC.element_to_be_clickable((By.XPATH, selector))
                        )
                        break
                    except:
                        continue
                
                if search_box:
                    search_box.click()
                    search_box.clear()
                    search_box.send_keys(self.group_name)
                    time.sleep(3)
                
                    # Try different group selectors
                    group_selectors = [
                        f"//span[@title='{self.group_name}']",
                        f"//span[contains(text(), '{self.group_name}')]",
                        f"//div[contains(@title, '{self.group_name}')]",
                        f"//span[contains(@title, '{self.group_name}')]"
                    ]
                    
                    for g_selector in group_selectors:
                        try:
                            group = self.driver.find_element(By.XPATH, g_selector)
                            group.click()
                            print(f"✅ AUTO-CLICKED: {self.group_name}")
                            time.sleep(2)
                            return True
                        except:
                            continue
                
                print(f"  Attempt {attempt+1} failed")
                
            except Exception as e:
                print(f"  Error: {e}")
            
            time.sleep(2)
        
        # 🛡️ MANUAL FALLBACK (100% works)
        print("\n🎯 MANUAL SELECT - 100% SUCCESS")
        print("📋 INSTRUCTIONS:")
        print("1. Use mouse to find your group")
        print("2. CLICK your group chat")
        print("3. Press ENTER here:")
        input()
        print("✅ Group manually selected - monitoring started!")
        return True

    def parse_complaint(self, text):
        """Parse complaint format"""
        lines = [l.strip() for l in text.split('\n') if l.strip()]
        if len(lines) < 4: return None
        
        data = {}
        data['Time'] = lines[0]
        data['Date'] = re.search(r'\d+-\d+-\d+', lines[1]).group() if re.search(r'\d+-\d+-\d+', lines[1]) else ''
        data['Branch'] = re.sub(r'Branch.*', '', lines[2], flags=re.I).strip()
        data['Complain'] = '\n'.join(lines[3:])[:200]
        
        # Auto-categorize
        complaint = data['Complain'].lower()
        if any(w in complaint for w in ['staff', 'late', 'employee']):
            data['Category'] = 'staff issue'
        elif any(w in complaint for w in ['food', 'quality', 'taste']):
            data['Category'] = 'quality issue'
        else:
            data['Category'] = 'other'
            
        data['Response (Y / N)'] = 'N'
        return data

    def update_excel(self, data):
        """Safe Excel update"""
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            row = ws.max_row + 1
            
            # Column mapping
            cols = {
                'Date': 1, 'Time': 2, 'Branch': 3, 'Category': 4, 
                'Complain': 7, 'Response (Y / N)': 8
            }
            
            for key, col in cols.items():
                if key in data:
                    ws.cell(row=row, column=col, value=data[key])
            
            wb.save(self.excel_path)
            print(f"✅ Saved row {row}")
            return True
        except Exception as e:
            print(f"❌ Excel error: {e}")
            return False

    def monitor(self, interval=30):
        """Main monitoring loop"""
        self.safe_start_chrome()
        if not self.find_group(): return
        
        print(f"🚀 Monitoring every {interval}s (Ctrl+C to stop)")
        
        try:
            while True:
                try:
                    messages = self.driver.find_elements(By.XPATH, "//div[@data-testid='msg-container']")
                    new_count = 0
                    
                    for msg in messages[-20:]:  # Last 20 messages
                        text = msg.text.strip()
                        if len(text) < 20: continue
                        
                        msg_id = hashlib.md5(text.encode()).hexdigest()
                        
                        if msg_id in self.processed: continue
                        
                        complaint = self.parse_complaint(text)
                        if complaint:
                            if self.update_excel(complaint):
                                new_count += 1
                                self.processed.add(msg_id)
                    
                    self.save_cache()
                    
                    if new_count:
                        print(f"📊 Found {new_count} new complaints")
                    else:
                        print(f"⏳ No new complaints")
                        
                except Exception as e:
                    print(f"⚠️ Check error: {e}")
                
                time.sleep(interval)
                
        except KeyboardInterrupt:
            print("\n🛑 Stopped by user")
        finally:
            self.driver.quit()

# ========== CONFIG & RUN ==========
excel_path = "C:\\Users\\GT-Tech\\Desktop\\Feb_surveillance.xlsx"  # Update this path
group = "Hns Surveillance"

bot = WhatsAppExcelUltimate(excel_path, group)
bot.monitor()


